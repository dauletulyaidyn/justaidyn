#!/usr/bin/env python3
import argparse
import csv
import json
import os
import sys
from datetime import date, datetime, timedelta


def normalize_key(key: str) -> str:
    return (key or "").strip().lower().replace(" ", "").replace("_", "")


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    return None


def compute_end_date(start_iso):
    if not start_iso:
        return None
    dt = datetime.strptime(start_iso, "%Y-%m-%d").date()
    # 4 недели + 1 день (включительно): +28 дней
    end_dt = dt + timedelta(days=28)
    return end_dt.isoformat()


def parse_int(value):
    if value is None:
        return 0
    text = str(value).strip()
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def map_row(raw):
    mapped = {}
    for key, value in raw.items():
        nk = normalize_key(key)
        mapped[nk] = value

    cohort_id = (
        mapped.get("id")
        or mapped.get("поток")
        or mapped.get("лек")
        or mapped.get("cohort")
        or ""
    )

    start_raw = (
        mapped.get("startdate")
        or mapped.get("началодаты")
        or mapped.get("басталукүні")
        or mapped.get("басталукүні")
        or mapped.get("басталуы")
        or mapped.get("start")
    )

    end_raw = (
        mapped.get("enddate")
        or mapped.get("окончаниедаты")
        or mapped.get("аяқталукүні")
        or mapped.get("аяқталуы")
        or mapped.get("end")
    )

    g10 = (
        mapped.get("group10")
        or mapped.get("топ10")
        or mapped.get("группа10")
        or 0
    )
    g50 = (
        mapped.get("group50")
        or mapped.get("топ50")
        or mapped.get("группа50")
        or 0
    )
    g50p = (
        mapped.get("group50+")
        or mapped.get("group50plus")
        or mapped.get("топ50+")
        or mapped.get("группа50+")
        or 0
    )

    start_date = parse_date(start_raw)
    end_date = parse_date(end_raw) or compute_end_date(start_date)

    return {
        "id": str(cohort_id).strip() or "—",
        "start_date": start_date,
        "end_date": end_date,
        "group10": parse_int(g10),
        "group50": parse_int(g50),
        "group50plus": parse_int(g50p),
    }


def read_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [map_row(r) for r in reader]


def read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Ошибка: для чтения .xlsx нужен openpyxl. Установите: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row_values in rows[1:]:
        if not any(v is not None and str(v).strip() for v in row_values):
            continue
        raw = {headers[i]: row_values[i] for i in range(min(len(headers), len(row_values)))}
        result.append(map_row(raw))
    return result


def main():
    parser = argparse.ArgumentParser(description="Импорт расписания курса из Excel/CSV в JSON")
    parser.add_argument("--input", required=True, help="Путь к файлу .xlsx или .csv")
    parser.add_argument("--output", default="data/course-schedule.json", help="Путь к итоговому JSON")
    args = parser.parse_args()

    in_path = args.input
    ext = os.path.splitext(in_path)[1].lower()

    if ext == ".csv":
        cohorts = read_csv(in_path)
    elif ext == ".xlsx":
        cohorts = read_xlsx(in_path)
    else:
        print("Ошибка: поддерживаются только .xlsx и .csv", file=sys.stderr)
        sys.exit(1)

    payload = {"cohorts": cohorts}

    out_dir = os.path.dirname(args.output)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    # Avoid console encoding issues on Windows terminals with non-UTF code pages.
    print(f"Done: wrote {len(cohorts)} rows to {args.output}")


if __name__ == "__main__":
    main()
